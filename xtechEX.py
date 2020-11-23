from bs4 import BeautifulSoup
import requests
import openpyxl as px
import os

if os.path.exists('HybridIT.xlsx'):
    wb = px.load_workbook('HybridIT.xlsx')
    sheetname = wb.sheetnames
    for name in sheetname:
        if 'xtech' in name:
            wb.remove(wb['xtech'])
    wb.create_sheet(title='xtech')
    ws = wb['xtech']
else:
    wb = px.Workbook()
    ws = wb.active
    ws.title = 'xtech'

title_data = dict()
titleList = list()
checklist = list()

res = requests.get('https://xtech.nikkei.com/')
res.raise_for_status()
soup = BeautifulSoup(res.text, "html.parser")
titles = soup.select('.c-linkArrowList_item' + ' a')

f = open('keywords.txt', 'r', encoding='utf-8')
keywords = f.readlines()
ks = len(keywords)
k1 = 1
kcn = 1

for title in titles:
    title_data = dict()
    title_data["title"] = title.getText()
    link = title.get('href')
    if 'https' in link:
        pass
    else:
        link ='https://xtech.nikkei.com/' + link
    title_data["URL"] = link
    title_data["content"] = None
    titleList.append(title_data)


for i in range(len(titleList)):
    link = titleList[i]["URL"]
    print(link)
    print(titleList[i]["title"])
    title = titleList[i]["title"]
    html = requests.get(link)
    soup = BeautifulSoup(html.text, "html.parser")
    
    contents = soup.select('p')
    for content in contents:
        if ks > 1:
            for keyword in keywords:
                keyword = keyword.strip()
                if keyword in content.text:
                    print("------------")
                    print(keyword)
                    print("------------")
                    ws.cell(row=1, column=kcn).value = 'キーワード: '+ keyword
                    if title in checklist:
                        pass
                    else:
                        ws.cell(row=k1+1, column=kcn).value = title
                        checklist.append(title)
                        ws.cell(row=k1+2, column=kcn).value = link
                        k1 +=3
                kcn += 1
            kcn = 1
        else:
            for keyword in keywords:
                keyword = keyword.strip()
                if keyword in content.text:
                    print("------------")
                    print(keyword)
                    print("------------")
                    ws.cell(row=1, column=1).value = 'キーワード: '+ keyword
                    if title in checklist:
                        pass
                    else:
                        ws.cell(row=k1+1, column=1).value = title
                        checklist.append(title)
                        ws.cell(row=k1+2, column=1).value = link
                        k1 +=3

wb.save('HybridIT.xlsx')

