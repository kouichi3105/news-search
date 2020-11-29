from bs4 import BeautifulSoup
import requests
import openpyxl as px
import os

title_data = dict()
titleList = list()
checklist = list()
proxy_data = dict()

if os.path.exists('HybridIT.xlsx'):
    wb = px.load_workbook('HybridIT.xlsx')
    sheetname = wb.sheetnames
    for name in sheetname:
        if 'ZDNET' in name:
            wb.remove(wb['ZDNET'])
    wb.create_sheet(title='ZDNET')
    ws = wb['ZDNET']
else:
    wb = px.Workbook()
    ws = wb.active
    ws.title = 'ZDNET'
ic = 1

if os.path.exists('proxy.txt'):
    f = open('proxy.txt', 'r', encoding='utf-8')
    proxies = f.readlines()
    if not proxies:
        proxy_data = None
    else:
        proxy_data["http"] = proxies[0].strip()
        proxy_data["https"] = proxies[1].strip()
else:
    proxy_data = None

if proxy_data is not None:
    res = requests.get('https://japan.zdnet.com/' , proxies=proxy_data)
else:
    res = requests.get('https://japan.zdnet.com/')

res.raise_for_status()
soup = BeautifulSoup(res.text, "html.parser")
titles = soup.select('.pg-container-main a')

f = open('keywords.txt', 'r', encoding='utf-8')
keywords = f.readlines()
ks = len(keywords)

k1 = 1
kcn = 1

for title in titles:
    title_data = dict()
    title_data["title"] = title.getText().strip().replace('\n', ' ').replace('\r', '')
    if title_data["title"] == '':
        title_data["title"] = None
    link = title.get('href')
    if 'https' in link:
        pass
    else:
        link ='https://japan.zdnet.com/' + link
    title_data["URL"] = link
    title_data["content"] = None
    titleList.append(title_data)

for i in range(len(titleList)):
    link = titleList[i]["URL"]
    print(link)
    print(titleList[i]["title"])
    title = titleList[i]["title"]
    html = requests.get(link)
    soup = BeautifulSoup(html.text, "html.parser")
    contents = soup.select('.article-contents')

    for content in contents:
        if ks > 1:
            for keyword in keywords:
                keyword = keyword.strip()
                if keyword in content.text:
                    print("------------")
                    print(keyword)
                    print("------------")
                    ws.cell(row=1, column=kcn).value = 'キーワード: '+ keyword
                    if title is not None:
                        ws.cell(row=k1+1, column=kcn).value = title
                        ws.cell(row=k1+2, column=kcn).value = link
                        k1 +=3
                    else:
                        k1 += 1
                kcn += 1
            kcn = 1
        else:
            for keyword in keywords:
                keyword = keyword.strip()
                if keyword in content.text:
                    print("------------")
                    print(keyword)
                    print("------------")
                    ws.cell(row=1, column=1).value = 'キーワード: '+ keyword
                    if title is not None:
                        ws.cell(row=k1+1, column=1).value = title
                        ws.cell(row=k1+2, column=1).value = link
                        k1 += 3
                    else:
                        k1 += 1
wb.save('HybridIT.xlsx')